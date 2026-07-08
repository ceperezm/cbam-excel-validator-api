from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.validators import EXPECTED_HEADERS


def ensure_xlsx_filename(filename: str | None) -> None:
    if not filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are allowed",
        )


def load_template_sheet(file_bytes: bytes) -> Worksheet:
    """
    Abre el libro Excel y selecciona la hoja Template o la primera hoja disponible.
    """
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 - converted into public API error
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file could not be read as a valid .xlsx workbook",
        ) from exc

    if "Template" in workbook.sheetnames:
        return workbook["Template"]
    return workbook[workbook.sheetnames[0]]


def normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_headers(sheet: Worksheet) -> list[str]:
    return [normalize_header(cell.value) for cell in sheet[1]][: len(EXPECTED_HEADERS)]


def validate_headers(sheet: Worksheet) -> None:
    """
    Verifica que los encabezados coincidan exactamente con el template esperado.
    """
    received_headers = read_headers(sheet)
    if received_headers != EXPECTED_HEADERS:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Invalid template headers",
                "expected_headers": EXPECTED_HEADERS,
                "received_headers": received_headers,
            },
        )


def row_is_empty(values: list[Any]) -> bool:
    return all(value is None or (isinstance(value, str) and value.strip() == "") for value in values)


def read_data_rows(file_bytes: bytes) -> list[tuple[int, dict[str, Any]]]:
    """
    Lee las filas de datos del template y conserva el numero real de fila de Excel.
    """
    sheet = load_template_sheet(file_bytes)
    validate_headers(sheet)

    rows: list[tuple[int, dict[str, Any]]] = []
    # Omite filas vacias para no generar errores innecesarios.
    for excel_row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(EXPECTED_HEADERS)])
        if row_is_empty(values):
            continue
        rows.append((excel_row_number, dict(zip(EXPECTED_HEADERS, values, strict=True))))
    return rows
